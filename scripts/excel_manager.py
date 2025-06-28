import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, timedelta

from stock_data_handler import StockDataHandler

class ExcelManager:
    # Counter for how many stocks updated
    _stocks_updated = 0


    def __init__(self, file_path):
        self.file_path = file_path

    def save_excel(self, df):
        try:
            # Load existing workbook to preserve formatting
            wb = load_workbook(self.file_path)
            ws = wb.active
            
            # Clear existing data (keep headers)
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).value = None
            
            # Write new data starting from row 2 (preserve headers)
            for index, row in df.iterrows():
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=index + 2, column=col_idx, value=value)
                    
                    # Format percentage columns
                    column_name = ws.cell(row=1, column=col_idx).value
                    if column_name and 'Change' in column_name and value is not None:
                        # Format as percentage with 2 decimal places
                        cell.number_format = '0.00%'
                    
                    # Format date columns
                    if column_name and 'Date' in column_name and value is not None:
                        # Format as date only (no time)
                        cell.number_format = 'yyyy-mm-dd'
            
            # Save the workbook
            wb.save(self.file_path)
            
        except PermissionError:
            print(f"Error: Cannot save to '{self.file_path}' - the file is currently open in Excel.")
            print("Please close the Excel file and try again.")
            return False
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            return False
        
        return True
        
    def create_excel(self):
        # Ensure file doesnt exist
        if os.path.exists(self.file_path):
            print(f"Error: File '{self.file_path}' already exists.")
            return

        # Define the column names for stock tracking
        columns = [
            'Symbol',           # Stock symbol (e.g., AAPL, MSFT)
            'Purchase Date',    # Date you bought/started tracking
            'Initial Price',    # Price when you started tracking

            # Summary
            'End Price',   # Final price after 7 days
            'Total Change Pct', # Total % change over the week
            
            # Daily tracking (7 days)
            'Day 1 Price',       # Price on day 1
            'Day 1 Change Pct',  # % change from initial price
            'Day 2 Price',       # Price on day 2
            'Day 2 Change Pct',  # % change from initial price
            'Day 3 Price',       # Price on day 3
            'Day 3 Change Pct',  # % change from initial price
            'Day 4 Price',       # Price on day 4
            'Day 4 Change Pct',  # % change from initial price
            'Day 5 Price',       # Price on day 5
            'Day 5 Change Pct',  # % change from initial price
            'Day 6 Price',       # Price on day 6
            'Day 6 Change Pct',  # % change from initial price
            'Day 7 Price',       # Price on day 7
            'Day 7 Change Pct',  # % change from initial price
        ]
        
        try:
        # Create empty DataFrame with the column structure
            df = pd.DataFrame(columns=columns)
            
            # Create a new workbook and select the active sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Stock Tracking"
            
            # Add the column headers
            for col, header in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col, value=header)
                
                # Format header cells
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, size=12)
            
            # Auto-adjust column widths based on content
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width (add some padding)
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Set row height for header
            # ws.row_dimensions[1].height = 25
            
            # Save the workbook
            wb.save(self.file_path)
            
            print(f"Successfully created {self.file_path}")
            print(f"Number of columns: {len(columns)}")
            print(f"Columns created with auto-width formatting:")
            for i, column in enumerate(columns, 1):
                print(f"  {i:2d}. {column}")
                
        except Exception as e:
            print(f"Error creating Excel file: {e}")

    def print_excel_columns(self):
        # Read the Excel into a pandas dataframe
        df = pd.read_excel(self.file_path)
        print(df.columns)

    def update_excel(self):
        # Read the Excel into a pandas dataframe
        df = pd.read_excel(self.file_path)
        # Convert data types to work
        df['Purchase Date'] = df['Purchase Date'].astype(str)
        # Print out the first 10 symbols
        print(f"\nFirst 10 symbols:")
        symbols = df['Symbol'].head(10)
        for i, symbol in enumerate(symbols, 1):
            print(f"  {i}. {symbol}")
        
        # Update the stock data for each row
        for row_index in range(len(df)):
            df = self.update_stock_data(df, row_index)

        # Save the changes back to Excel
        self.save_excel(df)
    
    def update_stock_data(self, df, row_index):
        self._stocks_updated += 1
        print(f"Updating stock #{self._stocks_updated}")
        # Update the stock data for a given row index
        # Check if "Day 7 Price" is not nan, can skip the update because it's already done
        if not pd.isna(df.at[row_index, 'Day 7 Price']):
            print(f"Skipping update for {df.at[row_index, 'Symbol']} because it's already done")
            return df

        # Get the symbol
        symbol = df.at[row_index, 'Symbol']

        # Purchase date
        purchase_date_str = df.at[row_index, 'Purchase Date']
        current_date_str = datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=-5))).strftime("%Y-%m-%d")
        # If no purchase date, get the current date at utc-5
        if purchase_date_str == 'nan':
            # Update the purchase date
            df.at[row_index, 'Purchase Date'] = current_date_str
            purchase_date_str = current_date_str

        # Ensure purchase_date_str is in correct format (no time component)
        purchase_date_str = purchase_date_str.split(' ')[0]
        purchase_date = datetime.strptime(purchase_date_str, "%Y-%m-%d")
        # print(f"Purchase date: {purchase_date}")
        
        # Save the clean date string back to the DataFrame
        df.at[row_index, 'Purchase Date'] = purchase_date_str
        
        current_date = datetime.strptime(current_date_str, "%Y-%m-%d")
        
        # Calculate the end date (7 days from purchase date)
        end_date = purchase_date + timedelta(days=6)  # 7 days total (purchase date + 6 more days)
        if end_date > current_date:
            end_date = current_date
        end_date_str = end_date.strftime("%Y-%m-%d")
        # Get the stock price data
        stock_data = StockDataHandler.get_stock_data_from_api(symbol, purchase_date_str, end_date_str)
        if not stock_data:
            print(f"No stock data found for {symbol}")
            return df
        # print(f"Stock data: {stock_data}")

        # Initial price
        initial_price = df.at[row_index, 'Initial Price']
        # If no initial price, get the current price
        if pd.isna(df.at[row_index, 'Initial Price']):
            # Get the current price
            # Ensure the date is in the stock data
            if purchase_date_str in stock_data:
                initial_price = stock_data[purchase_date_str]['open']
            else:
                # Use first date in stock data
                initial_price = stock_data[list(stock_data.keys())[0]]['open']

            # Update the initial price
            df.at[row_index, 'Initial Price'] = initial_price
        end_price = initial_price
        
        # Calculate days between purchase date and current date
        days_diff = (end_date - purchase_date).days
        
        print(f"Symbol: {symbol}, Purchase Date: {purchase_date}, Initial Price: {initial_price}")
        print(f"Tracking period: {purchase_date} to {end_date_str}")
        print(f"Days since purchase: {days_diff}")
        
        # Iterate through dates from purchase date forwards for 7 days
        days_to_track = days_diff + 1
        
        for day_num in range(1, days_to_track + 1): # +1 because python doesn't iterate the last number
            # Calculate the date for this day (going forwards from purchase date)
            target_date = purchase_date + timedelta(days=day_num-1)
            target_date_str = target_date.strftime("%Y-%m-%d")
            
            # print(f"  Day {day_num}: {target_date_str}")
            
            # Check if this day has already happened (target_date <= current_date)
            if target_date <= current_date:
                # Update the price for this day (using day number for now)
                # See if the date is in the stock data
                if target_date_str in stock_data:
                    price = stock_data[target_date_str]['close']
                else:
                    price = -1
                    # print(f"    Day {day_num} price not found in stock data")
                # If price is -1, set the price to the previous day's price
                if price == -1:
                    # First day, no previous day's price
                    if day_num == 1:
                        price = initial_price
                    # Prev days price
                    else:
                        price = df.at[row_index, f'Day {day_num - 1} Price']
                # Update the price for this day
                df.at[row_index, f'Day {day_num} Price'] = price
                # Calculate the change percentage
                change_pct = ((price - initial_price) / initial_price)
                df.at[row_index, f'Day {day_num} Change Pct'] = change_pct
                # Update the end price so its always the latest price
                end_price = price


                # Calculate percentage change from initial price
                # if initial_price and initial_price > 0:
                    # change_pct = ((day_num - initial_price) / initial_price) * 100
                    # df.at[row_index, f'Day {day_num} Change Pct'] = change_pct
                
                # print(f"    Updated Day {day_num} Price to {price}")
            else:
                # print(f"    Day {day_num} hasn't happened yet")
                pass
        
        # Update the end price and change percentage
        df.at[row_index, 'End Price'] = end_price
        df.at[row_index, 'Total Change Pct'] = ((end_price - initial_price) / initial_price)

        return df

