"""
Excel Service Module

This module provides services for managing Excel files containing stock tracking data.
It handles reading, writing, and updating Excel files with proper formatting.
"""

import pandas as pd
import os
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, timedelta
from typing import Dict, Any, Optional

from .stock_data_service import StockDataService

# Configure logging for this module
logger = logging.getLogger(__name__)


class ExcelService:
    """
    Service class for managing Excel files with stock tracking data.
    
    This class provides methods to create, read, update, and save Excel files
    with proper formatting and data validation.
    """
    
    # Counter for tracking how many stocks have been updated
    _stocks_updated = 0
    
    # Column definitions for stock tracking
    STOCK_TRACKING_COLUMNS = [
        'Symbol',           # Stock symbol (e.g., AAPL, MSFT)
        'Purchase Date',    # Date you bought/started tracking
        'Initial Price',    # Price when you started tracking
        
        # Summary
        'End Price',        # Final price after 7 days
        'Total Change Pct', # Total % change over the week
        
        # Daily tracking (7 days)
        'Day 1 Price',       # Price on day 1
        'Day 1 Change Pct',  # % change from initial price
        'Day 2 Price',       # Price on day 2
        'Day 2 Change Pct',  # % change from initial price
        'Day 3 Price',       # Price on day 3
        'Day 3 Change Pct',  # % change from initial price
        'Day 4 Price',       # Price on day 4
        'Day 4 Change Pct',  # % change from initial price
        'Day 5 Price',       # Price on day 5
        'Day 5 Change Pct',  # % change from initial price
        'Day 6 Price',       # Price on day 6
        'Day 6 Change Pct',  # % change from initial price
        'Day 7 Price',       # Price on day 7
        'Day 7 Change Pct',  # % change from initial price
    ]

    def __init__(self, file_path: str):
        """
        Initialize the Excel service with a file path.
        
        Args:
            file_path: Path to the Excel file to manage
        """
        self.file_path = file_path
        logger.info(f"Initialized ExcelService for file: {file_path}")

    def save_excel(self, dataframe: pd.DataFrame) -> bool:
        """
        Save a pandas DataFrame to Excel with proper formatting.
        
        Args:
            dataframe: DataFrame to save to Excel
            
        Returns:
            True if successful, False otherwise
        """
        try:
            logger.info(f"Saving Excel file: {self.file_path} with {len(dataframe)} rows")
            
            # Load existing workbook to preserve formatting
            workbook = load_workbook(self.file_path)
            worksheet = workbook.active
            
            # Clear existing data (keep headers)
            for row in range(2, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=row, column=col).value = None
            
            # Write new data starting from row 2 (preserve headers)
            for index, row_data in dataframe.iterrows():
                for col_idx, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=index + 2, column=col_idx, value=value)
                    
                    # Format percentage columns
                    column_name = worksheet.cell(row=1, column=col_idx).value
                    if column_name and 'Change' in column_name and value is not None:
                        # Format as percentage with 2 decimal places
                        cell.number_format = '0.00%'
                    
                    # Format date columns
                    if column_name and 'Date' in column_name and value is not None:
                        # Format as date only (no time)
                        cell.number_format = 'yyyy-mm-dd'
            
            # Save the workbook
            workbook.save(self.file_path)
            logger.info(f"Successfully saved Excel file: {self.file_path}")
            return True
            
        except PermissionError:
            logger.error(f"Permission error saving Excel file: {self.file_path} - file may be open in Excel")
            print(f"Error: Cannot save to '{self.file_path}' - the file is currently open in Excel.")
            print("Please close the Excel file and try again.")
            return False
        except Exception as error:
            logger.error(f"Error saving Excel file {self.file_path}: {error}")
            print(f"Error saving Excel file: {error}")
            return False

    def create_excel(self) -> None:
        """
        Create a new Excel file with the proper structure for stock tracking.
        
        Creates a workbook with predefined columns and formatting for tracking
        stock performance over a 7-day period.
        """
        # Ensure file doesn't exist
        if os.path.exists(self.file_path):
            logger.warning(f"Excel file already exists: {self.file_path}")
            print(f"Error: File '{self.file_path}' already exists.")
            return

        try:
            logger.info(f"Creating new Excel file: {self.file_path}")
            
            # Create empty DataFrame with the column structure
            dataframe = pd.DataFrame(columns=self.STOCK_TRACKING_COLUMNS)
            
            # Create a new workbook and select the active sheet
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Stock Tracking"
            
            # Add the column headers
            for col, header in enumerate(self.STOCK_TRACKING_COLUMNS, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                
                # Format header cells
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, size=12)
            
            # Auto-adjust column widths based on content
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width (add some padding)
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            workbook.save(self.file_path)
            
            logger.info(f"Successfully created Excel file: {self.file_path} with {len(self.STOCK_TRACKING_COLUMNS)} columns")
            print(f"Successfully created {self.file_path}")
            print(f"Number of columns: {len(self.STOCK_TRACKING_COLUMNS)}")
            print("Columns created with auto-width formatting:")
            for i, column in enumerate(self.STOCK_TRACKING_COLUMNS, 1):
                print(f"  {i:2d}. {column}")
                
        except Exception as error:
            logger.error(f"Error creating Excel file {self.file_path}: {error}")
            print(f"Error creating Excel file: {error}")

    def print_excel_columns(self) -> None:
        """
        Print the column names from the Excel file.
        
        Reads the Excel file and displays all column names for verification.
        """
        try:
            logger.info(f"Reading Excel columns from: {self.file_path}")
            
            # Read the Excel into a pandas dataframe
            dataframe = pd.read_excel(self.file_path)
            print("Excel columns:")
            for i, column in enumerate(dataframe.columns, 1):
                print(f"  {i:2d}. {column}")
                
            logger.info(f"Found {len(dataframe.columns)} columns in Excel file")
            
        except Exception as error:
            logger.error(f"Error reading Excel columns from {self.file_path}: {error}")
            print(f"Error reading Excel columns: {error}")

    def update_excel(self) -> None:
        """
        Update the Excel file with current stock data.
        
        Reads the existing Excel file, fetches current stock data for each symbol,
        and saves the updated data back to the file.
        """
        try:
            logger.info(f"Starting Excel update process for: {self.file_path}")
            
            # Read the Excel into a pandas dataframe
            dataframe = pd.read_excel(self.file_path)
            logger.info(f"Loaded Excel file with {len(dataframe)} rows")
            
            # Convert data types to work properly
            dataframe['Purchase Date'] = dataframe['Purchase Date'].astype(str)
            
            # Print out the first 10 symbols
            print(f"\nFirst 10 symbols:")
            symbols = dataframe['Symbol'].head(10)
            for i, symbol in enumerate(symbols, 1):
                print(f"  {i}. {symbol}")
            
            logger.info(f"Processing {len(dataframe)} stocks for update")
            
            # Update the stock data for each row
            for row_index in range(len(dataframe)):
                dataframe = self._update_stock_data(dataframe, row_index)

            # Save the changes back to Excel
            self.save_excel(dataframe)
            
            logger.info(f"Completed Excel update process for: {self.file_path}")
            
        except Exception as error:
            logger.error(f"Error updating Excel file {self.file_path}: {error}")
            print(f"Error updating Excel file: {error}")
    
    def _update_stock_data(self, dataframe: pd.DataFrame, row_index: int) -> pd.DataFrame:
        """
        Update stock data for a specific row in the DataFrame.
        
        Args:
            dataframe: DataFrame containing stock data
            row_index: Index of the row to update
            
        Returns:
            Updated DataFrame
        """
        self._stocks_updated += 1
        symbol = dataframe.at[row_index, 'Symbol']
        
        logger.info(f"Updating stock #{self._stocks_updated}: {symbol}")
        print(f"Updating stock #{self._stocks_updated}")
        
        # Check if "Day 7 Price" is not nan, can skip the update because it's already done
        if not pd.isna(dataframe.at[row_index, 'Day 7 Price']):
            logger.info(f"Skipping update for {symbol} - already completed")
            print(f"Skipping update for {symbol} because it's already done")
            return dataframe

        # Purchase date
        purchase_date_str = dataframe.at[row_index, 'Purchase Date']
        current_date_str = datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=-5))).strftime("%Y-%m-%d")
        
        # If no purchase date, get the current date at utc-5
        if purchase_date_str == 'nan':
            # Update the purchase date
            dataframe.at[row_index, 'Purchase Date'] = current_date_str
            purchase_date_str = current_date_str
            logger.info(f"Set purchase date for {symbol} to: {current_date_str}")

        # Ensure purchase_date_str is in correct format (no time component)
        purchase_date_str = purchase_date_str.split(' ')[0]
        purchase_date = datetime.strptime(purchase_date_str, "%Y-%m-%d")
        
        # Save the clean date string back to the DataFrame
        dataframe.at[row_index, 'Purchase Date'] = purchase_date_str
        
        current_date = datetime.strptime(current_date_str, "%Y-%m-%d")
        
        # Calculate the end date (7 days from purchase date)
        end_date = purchase_date + timedelta(days=6)  # 7 days total (purchase date + 6 more days)
        if end_date > current_date:
            end_date = current_date
        end_date_str = end_date.strftime("%Y-%m-%d")
        
        logger.info(f"Fetching data for {symbol} from {purchase_date_str} to {end_date_str}")
        
        # Get the stock price data
        stock_data = StockDataService.get_stock_data_from_api(symbol, purchase_date_str, end_date_str)
        
        if not stock_data:
            logger.warning(f"No data found for {symbol}")
            print(f"No data found for {symbol}")
            return dataframe

        # Update the DataFrame with the stock data
        dataframe = self._populate_stock_data(dataframe, row_index, stock_data, purchase_date_str)
        
        logger.info(f"Successfully updated data for {symbol}")
        return dataframe

    def _populate_stock_data(self, dataframe: pd.DataFrame, row_index: int, 
                           stock_data: Dict[str, Dict[str, Any]], purchase_date_str: str) -> pd.DataFrame:
        """
        Populate the DataFrame with stock data for a specific row.
        
        Args:
            dataframe: DataFrame to update
            row_index: Index of the row to update
            stock_data: Stock data dictionary
            purchase_date_str: Purchase date string
            
        Returns:
            Updated DataFrame
        """
        symbol = dataframe.at[row_index, 'Symbol']
        logger.info(f"Populating stock data for {symbol} with {len(stock_data)} data points")
        
        # Get sorted dates from stock data
        sorted_dates = sorted(stock_data.keys())
        
        if not sorted_dates:
            logger.warning(f"No sorted dates available for {symbol}")
            return dataframe
        
        # Set initial price (open price from first available day)
        first_day_data = stock_data[sorted_dates[0]]
        initial_price = first_day_data.get('open')
        
        # If open price is not available for the first day, try to find the first day with open data
        if initial_price is None:
            logger.warning(f"No open price available for {symbol} on first day {sorted_dates[0]}, searching for first day with open data")
            for date in sorted_dates:
                day_data = stock_data[date]
                if day_data.get('open') is not None:
                    initial_price = day_data['open']
                    logger.info(f"Using open price from {date}: ${initial_price:.2f}")
                    break
        
        # If still no open price found, fall back to close price from first day
        if initial_price is None:
            initial_price = first_day_data.get('close')
            logger.warning(f"No open price found for {symbol}, falling back to close price from first day: ${initial_price:.2f}")
        
        dataframe.at[row_index, 'Initial Price'] = initial_price
        
        # Set end price (close price from last available day)
        end_price = stock_data[sorted_dates[-1]]['close']
        dataframe.at[row_index, 'End Price'] = end_price
        
        # Calculate total change percentage
        if initial_price and end_price:
            total_change_pct = (end_price - initial_price) / initial_price
            dataframe.at[row_index, 'Total Change Pct'] = total_change_pct
            logger.info(f"{symbol}: Initial=${initial_price:.2f} (open), End=${end_price:.2f} (close), Change={total_change_pct:.2%}")
        
        # Populate daily data (up to 7 days)
        for day_num in range(1, 8):
            if day_num - 1 < len(sorted_dates):
                date = sorted_dates[day_num - 1]
                price_data = stock_data[date]
                
                # Set day price (use close price for daily tracking)
                day_price = price_data['close']
                dataframe.at[row_index, f'Day {day_num} Price'] = day_price
                
                # Calculate and set day change percentage
                if initial_price and day_price:
                    day_change_pct = (day_price - initial_price) / initial_price
                    dataframe.at[row_index, f'Day {day_num} Change Pct'] = day_change_pct
        
        logger.info(f"Successfully populated data for {symbol}: {len(sorted_dates)} days processed")
        return dataframe


# Example usage
if __name__ == "__main__":
    # Example usage
    excel_file = "stock_tracking.xlsx"
    excel_service = ExcelService(excel_file)
    
    # Create a new Excel file
    excel_service.create_excel()
    
    # Print columns
    excel_service.print_excel_columns() 